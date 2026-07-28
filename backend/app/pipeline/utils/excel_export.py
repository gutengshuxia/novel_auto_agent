"""Excel 导出 —— 将多版本 Prompt + 分镜数据格式化为美观表格。

表格分两个 Sheet:
1. Storyboard 概览 —— 每个镜头的元数据(编号、时长、景别、运镜、镜头描述)
2. Prompt Variants —— 每个镜头 × 多模型(Sora / Runway / Kling)的展开 Prompt
"""

from __future__ import annotations

import os
from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from ..schemas.prompt_plan import PromptPlan
    from ..schemas.story_analysis import StoryAnalysis
    from ..schemas.storyboard import Storyboard

from .logger import get_logger

logger = get_logger(__name__)

# ----- 样式常量 -----
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

BODY_FONT = Font(name="微软雅黑", size=10)
BODY_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)

ALT_FILL = PatternFill("solid", fgColor="F2F6FA")

THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _autosize(ws, max_width: int = 60) -> None:
    """根据内容粗略估算列宽,避免内容被截断。"""
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 8
        for cell in column_cells:
            if cell.value is None:
                continue
            # 中文字符按 2 算宽度
            length = sum(2 if ord(ch) > 127 else 1 for ch in str(cell.value))
            if length > max_len:
                max_len = length
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, max_width)


def _write_header(ws, headers: list[str]) -> None:
    for col_idx, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER
    ws.row_dimensions[1].height = 28


def export_prompts_to_excel(
    storyboard: "Storyboard",
    prompt_plan: "PromptPlan",
    output_dir: str | os.PathLike[str] = "./output",
    story_title: str = "Untitled",
    story_analysis: "StoryAnalysis | None" = None,
    director_ids: list[str] | None = None,
    cast_data: dict[str, Any] | None = None,
    storyboard_cards: list[dict[str, Any]] | None = None,
) -> Path:
    """导出分镜与多版本 Prompt 至 Excel 文件,返回生成的文件路径。"""
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_title = "".join(c if c.isalnum() else "_" for c in story_title)[:40]
    file_path = out_dir / f"storyboard_{safe_title}_{timestamp}.xlsx"

    wb = Workbook()

    # ====== Sheet 1: Storyboard 概览 ======
    ws1 = wb.active
    ws1.title = "Storyboard"
    headers1 = [
        "镜头编号", "时长(s)", "景别", "运镜", "镜头描述", "出场角色", "关联台词"
    ]
    _write_header(ws1, headers1)

    # 把 shot.id -> dialogue 建立索引,便于 Sheet1 引用
    dialogue_map: dict[str, list[str]] = {}
    for sp in prompt_plan.shot_prompts:
        # dialogue 可能是 DialogueLine 对象列表, 需转为可读字符串
        raw_dialogue = sp.dialogue or []
        dialogue_map[sp.shot_id] = [
            f"[{d.character_id}] {d.line} ({d.emotion})" if hasattr(d, 'line') else str(d)
            for d in raw_dialogue
        ]

    for row_idx, shot in enumerate(storyboard.shots, start=2):
        values = [
            shot.shot_id,
            shot.duration_sec,
            shot.framing,
            shot.camera.value,
            shot.description,
            ", ".join(shot.characters_in_shot),
            "\n".join(dialogue_map.get(shot.shot_id, [])),
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.alignment = BODY_ALIGN
            cell.border = BORDER
            if row_idx % 2 == 0:
                cell.fill = ALT_FILL
        ws1.row_dimensions[row_idx].height = 60

    _autosize(ws1)

    # ====== Sheet 2: Prompt Variants ======
    ws2 = wb.create_sheet("Prompt Variants")
    headers2 = ["镜头编号", "目标模型", "导演风格", "Prompt 文本", "镜头描述(参考)", "备注"]
    _write_header(ws2, headers2)

    # 导演风格字符串
    director_style_str = ", ".join(director_ids) if director_ids else "默认"

    shot_desc_map = {s.shot_id: s.description for s in storyboard.shots}
    row_idx = 2
    for sp in prompt_plan.shot_prompts:
        for variant in sp.variants:
            values = [
                sp.shot_id,
                variant.target_model.value,
                director_style_str,
                variant.prompt_text,
                shot_desc_map.get(sp.shot_id, ""),
                variant.notes or "",
            ]
            for col_idx, val in enumerate(values, start=1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=val)
                cell.font = BODY_FONT
                cell.alignment = BODY_ALIGN
                cell.border = BORDER
                if row_idx % 2 == 0:
                    cell.fill = ALT_FILL
            ws2.row_dimensions[row_idx].height = 90
            row_idx += 1

    _autosize(ws2, max_width=80)

    # ====== Sheet 3: 演员表 (Character Sheets) ======
    ws3 = wb.create_sheet("演员表")
    headers3 = ["角色名", "角色定位", "演员表描述", "首次出现", "服装变化", "参考图URL"]
    _write_header(ws3, headers3)

    # 优先使用全局演员表 (cast_data), 否则使用 story_analysis
    if cast_data:
        row_idx = 2
        for name, data in cast_data.items():
            costumes = data.get("costumes", {})
            costumes_str = "; ".join([f"{ch}: {co}" for ch, co in costumes.items()]) if costumes else ""
            values = [
                name,
                data.get("role", ""),
                data.get("character_sheet", "") or data.get("base_appearance", ""),
                data.get("first_chapter", ""),
                costumes_str,
                data.get("reference_image_url", ""),
            ]
            for col_idx, val in enumerate(values, start=1):
                cell = ws3.cell(row=row_idx, column=col_idx, value=val)
                cell.font = BODY_FONT
                cell.alignment = BODY_ALIGN
                cell.border = BORDER
                if row_idx % 2 == 0:
                    cell.fill = ALT_FILL
            ws3.row_dimensions[row_idx].height = 120
            row_idx += 1
    elif story_analysis:
        row_idx = 2
        for ch in story_analysis.characters:
            values = [
                ch.name,
                ch.role or "",
                ch.character_sheet or ch.visual_anchor or ch.appearance or "",
                story_title,  # 首次出现章节
                "",  # 服装变化
                ch.reference_image_url or "",
            ]
            for col_idx, val in enumerate(values, start=1):
                cell = ws3.cell(row=row_idx, column=col_idx, value=val)
                cell.font = BODY_FONT
                cell.alignment = BODY_ALIGN
                cell.border = BORDER
                if row_idx % 2 == 0:
                    cell.fill = ALT_FILL
            ws3.row_dimensions[row_idx].height = 120
            row_idx += 1

    _autosize(ws3, max_width=80)
    ws3.freeze_panes = "A2"

    # ---- Sheet 4: 故事板分镜卡片 (Storyboard Cards) ----
    if storyboard_cards:
        ws4 = wb.create_sheet("故事板卡片")
        ws4.append(["镜头编号", "卡片类型", "卡片标题", "卡片提示词", "图片 URL"])
        for cell in ws4[1]:
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGN
            cell.fill = HEADER_FILL
            cell.border = BORDER

        row_idx = 2
        for card in storyboard_cards:
            values = [
                card.get("shot_id", ""),
                card.get("type", ""),
                card.get("title", ""),
                card.get("prompt", ""),
                card.get("image_url", "") or "",
            ]
            for col_idx, val in enumerate(values, start=1):
                cell = ws4.cell(row=row_idx, column=col_idx, value=val)
                cell.font = BODY_FONT
                cell.alignment = BODY_ALIGN
                cell.border = BORDER
                if row_idx % 2 == 0:
                    cell.fill = ALT_FILL
            ws4.row_dimensions[row_idx].height = 150
            row_idx += 1

        _autosize(ws4, max_width=80)
        ws4.freeze_panes = "A2"

    # 冻结首行
    ws1.freeze_panes = "A2"
    ws2.freeze_panes = "A2"

    wb.save(file_path)
    logger.info("Excel 已导出: %s", file_path)
    return file_path


__all__ = ["export_prompts_to_excel"]
