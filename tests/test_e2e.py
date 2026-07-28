"""E2E 测试 —— Mock LLM + Mock Pydantic + Mock LangGraph,在沙箱里跑通整图。

核心策略:
1. sys.modules 注入: pydantic / langchain_core.messages / langchain_openai 全部用轻量 mock
2. langgraph.graph.END / StateGraph 用一个迷你实现替代,支持 add_node / add_edge / add_conditional_edges / compile / invoke
3. ChatOpenAI 被 monkey-patch 成 "按调用顺序返回预设 JSON" 的 fake LLM
4. 真实运行 src.agents / src.graph / src.schemas / src.utils 的全部代码

测试场景:
- Scenario A: 顺畅通关 —— Step5 一次通过, 6 个节点走完, Excel 落盘
- Scenario B: 回滚重试 —— Step5 第一次 failed, 回滚 Step3, 第二次通过
- Scenario C (bonus): 回滚超限 —— Step5 连续 failed, 达到 max_replans 后终止

使用:
    python3 tests/test_e2e.py
"""
from __future__ import annotations

import json
import sys
import types
import pathlib

ROOT = pathlib.Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

# ============================================================
# Part 0: 沙箱 Mock 注入 (在 import 任何项目模块前完成)
# ============================================================

# --- 0.1 Mock pydantic ---
class _FieldInfo:
    def __init__(self, default=..., default_factory=None, description="", **kw):
        self.default = default
        self.default_factory = default_factory
        self.description = description
        for k, v in kw.items():
            setattr(self, k, v)


def Field(default=..., **kw):
    if "default_factory" in kw and default is ...:
        return _FieldInfo(default=None, default_factory=kw.pop("default_factory"), **kw)
    return _FieldInfo(default=default, default_factory=None, **kw)


class ValidationError(Exception):
    def __init__(self, errors):
        self.errors = errors if isinstance(errors, list) else [errors]
        super().__init__(f"ValidationError: {self.errors[:3]}")




def _find_nested_model(ftype, data):
    """根据类型注解字符串找到 BaseModel 子类。"""
    if not isinstance(ftype, str):
        return None
    target = ftype.strip()
    for subcls in BaseModel.__subclasses__():
        if subcls.__name__ == target:
            return subcls
    return None


def _find_list_item_cls(ftype):
    """list[X] -> X 类。"""
    if not isinstance(ftype, str):
        return None
    t = ftype.strip()
    if t.startswith("list[") and t.endswith("]"):
        inner = t[5:-1]
        for subcls in BaseModel.__subclasses__():
            if subcls.__name__ == inner:
                return subcls
    return None

class _ModelMetaclass(type):
    def __new__(mcs, name, bases, namespace):
        # 收集字段定义
        annotations = namespace.get("__annotations__", {})
        fields = {}
        for fname, ftype in annotations.items():
            if fname in namespace:
                fobj = namespace[fname]
                if isinstance(fobj, _FieldInfo):
                    fields[fname] = (ftype, fobj)
                else:
                    fields[fname] = (ftype, _FieldInfo(default=fobj))
            else:
                fields[fname] = (ftype, _FieldInfo())
        namespace["__fields__"] = fields
        return super().__new__(mcs, name, bases, namespace)






def _coerce_list_item(item_cls, item):
    if isinstance(item, dict):
        return item_cls.model_validate(item)
    # 枚举列表项
    if isinstance(item, str) and hasattr(item_cls, "__members__"):
        return item_cls(item)
    return item


# 在 mock 注入阶段就定义好所有枚举, 供 _coerce_value 查找
import enum as _enum

class TargetModel(str, _enum.Enum):
    KLING = "kling"
    JIMENG = "jimeng"

class FramingStyle(str, _enum.Enum):
    EXTREME_WIDE = "extreme_wide"
    WIDE = "wide"
    FULL = "full"
    MEDIUM_WIDE = "medium_wide"
    MEDIUM = "medium"
    MEDIUM_CLOSE = "medium_close"
    CLOSE_UP = "close_up"
    EXTREME_CLOSE_UP = "extreme_close_up"

class VisualStyle(str, _enum.Enum):
    CINEMATIC = "cinematic"
    ANIME = "anime"
    REALISTIC = "realistic"
    OIL_PAINTING = "oil_painting"
    WATERCOLOR = "watercolor"
    PIXEL_ART = "pixel_art"
    NOIR = "noir"
    CYBERPUNK = "cyberpunk"
    FANTASY = "fantasy"
    DOCUMENTARY = "documentary"

class MoodTone(str, _enum.Enum):
    DARK = "dark"
    HOPEFUL = "hopeful"
    TENSE = "tense"
    MYSTERIOUS = "mysterious"
    EPIC = "epic"
    WHIMSICAL = "whimsical"
    MELANCHOLIC = "melancholic"
    ROMANTIC = "romantic"
    HORROR = "horror"
    NEUTRAL = "neutral"

class AspectRatio(str, _enum.Enum):
    RATIO_16_9 = "16:9"
    RATIO_9_16 = "9:16"
    RATIO_1_1 = "1:1"
    RATIO_21_9 = "21:9"
    RATIO_4_3 = "4:3"

class CameraMovement(str, _enum.Enum):
    STATIC = "static"
    PAN_LEFT = "pan_left"; PAN_RIGHT = "pan_right"
    TILT_UP = "tilt_up"; TILT_DOWN = "tilt_down"
    DOLLY_IN = "dolly_in"; DOLLY_OUT = "dolly_out"
    TRACK_LEFT = "track_left"; TRACK_RIGHT = "track_right"
    CRANE_UP = "crane_up"; CRANE_DOWN = "crane_down"
    ZOOM_IN = "zoom_in"; ZOOM_OUT = "zoom_out"
    HANDHELD = "handheld"; DRONE_AERIAL = "drone_aerial"

# DeliveryType 在源码里是 Literal, 不需要枚举


_ENUM_REGISTRY = {e.__name__: e for e in [
    TargetModel, FramingStyle, VisualStyle, MoodTone, AspectRatio, CameraMovement
]}


def _find_enum_cls(ftype):
    if not isinstance(ftype, str):
        return None
    return _ENUM_REGISTRY.get(ftype.strip())



def _strip_list(ftype):
    if isinstance(ftype, str) and ftype.startswith("list[") and ftype.endswith("]"):
        return ftype[5:-1]
    return ftype
def _to_jsonable(v):
    """递归: BaseModel -> dict, Enum -> value, list/dict 递归。"""
    import enum as _e
    if isinstance(v, _e.Enum):
        return v.value
    if isinstance(v, BaseModel):
        return v.model_dump()
    if isinstance(v, list):
        return [_to_jsonable(x) for x in v]
    if isinstance(v, dict):
        return {k: _to_jsonable(x) for k, x in v.items()}
    return v

class BaseModel(metaclass=_ModelMetaclass):
    def __init__(self, **data):
        # 收集: 既要尊重传入值,也要尊重字段 default
        for fname, (ftype, finfo) in self.__class__.__fields__.items():
            if fname in data:
                value = data[fname]
                if isinstance(ftype, str) and ftype.startswith("list[") and value is None:
                    value = []
                setattr(self, fname, value)
            else:
                if finfo.default is not ...:
                    setattr(self, fname, finfo.default)
                elif finfo.default_factory is not None:
                    setattr(self, fname, finfo.default_factory())
                else:
                    raise ValidationError([{"loc": (fname,), "msg": "missing"}])

    @classmethod
    def model_validate(cls, data):
        if not isinstance(data, dict):
            raise ValidationError([{"msg": f"expected dict, got {type(data).__name__}"}])
        # 递归: 把 dict 转为嵌套 BaseModel, list 中的 dict 也要转
        coerced = cls._coerce(data)
        return cls(**coerced)



    @classmethod
    def _coerce_value(cls, v, fname):
        ftype = cls.__fields__.get(fname, (None, None))[0]
        # 枚举: 字段是 Enum 子类, v 是字符串
        enum_cls = _find_enum_cls(ftype)
        if enum_cls and isinstance(v, str):
            return enum_cls(v)
        if isinstance(v, dict):
            nested_cls = _find_nested_model(ftype, v)
            if nested_cls:
                return nested_cls.model_validate(v)
            return v
        if isinstance(v, list):
            item_cls = _find_list_item_cls(ftype)
            enum_item_cls = _find_enum_cls(_strip_list(ftype))
            if item_cls:
                return [_coerce_list_item(item_cls, item) for item in v]
            if enum_item_cls:
                return [enum_item_cls(item) for item in v]
            return v
        return v

    @classmethod
    def _coerce(cls, data):
        if not isinstance(data, dict):
            return data
        out = {}
        for fname, (ftype, _finfo) in cls.__fields__.items():
            if fname not in data:
                continue
            out[fname] = cls._coerce_value(data[fname], fname)
        return out

    def model_dump(self, exclude_none=False, **kw):
        out = {}
        for fname in self.__class__.__fields__:
            if not hasattr(self, fname):
                continue
            v = getattr(self, fname)
            if exclude_none and v is None:
                continue
            out[fname] = _to_jsonable(v)
        return out

    def model_dump_json(self, indent=None, exclude_none=False, **kw):
        return json.dumps(self.model_dump(exclude_none=exclude_none),
                          ensure_ascii=False, indent=indent)

    @classmethod
    def model_json_schema(cls):
        props = {}
        required = []
        for fname, (ftype, finfo) in cls.__fields__.items():
            props[fname] = {"type": "string", "description": finfo.description}
            if finfo.default is ...:
                required.append(fname)
        return {
            "title": cls.__name__,
            "type": "object",
            "properties": props,
            "required": required,
        }


def field_validator(*fields):
    """装饰器: 记录但不实际校验。"""
    def deco(fn):
        fn.__fields_validated__ = fields
        return fn
    return deco


def model_validator(mode="after"):
    def deco(fn):
        fn.__model_validator_mode__ = mode
        return fn
    return deco


pyd_mod = types.ModuleType("pydantic")
pyd_mod.BaseModel = BaseModel
pyd_mod.Field = Field
pyd_mod.ValidationError = ValidationError
pyd_mod.field_validator = field_validator
pyd_mod.model_validator = model_validator
sys.modules["pydantic"] = pyd_mod

# 0.2 Mock langchain_core.messages
class _Msg:
    def __init__(self, content=""):
        self.content = content

class AIMessage(_Msg): pass
class HumanMessage(_Msg): pass
class SystemMessage(_Msg): pass

lc_messages = types.ModuleType("langchain_core.messages")
lc_messages.AIMessage = AIMessage
lc_messages.HumanMessage = HumanMessage
lc_messages.SystemMessage = SystemMessage
sys.modules["langchain_core"] = types.ModuleType("langchain_core")
sys.modules["langchain_core.messages"] = lc_messages

# 0.3 Mock langchain_openai.ChatOpenAI
class _FakeChatOpenAI:
    """占位类, 真正的 fake 由 monkey-patch get_llm 注入。"""
    def __init__(self, *a, **kw): pass
    def invoke(self, messages):
        raise RuntimeError("FakeChatOpenAI.invoke should be patched")


lc_openai = types.ModuleType("langchain_openai")
lc_openai.ChatOpenAI = _FakeChatOpenAI
sys.modules["langchain_openai"] = lc_openai

# 0.4 Mock langgraph.graph (迷你实现)
END = "__END__"

class _MiniGraph:
    def __init__(self, nodes, edges, conditional_edges, entry_point):
        self.nodes = nodes
        self.edges = edges                # {(from,): to}
        self.conditional_edges = conditional_edges  # {(from,): (router_fn, mapping)}
        self.entry_point = entry_point

    def invoke(self, initial_state):
        state = dict(initial_state)
        trajectory = []
        current = self.entry_point

        # 最多循环 100 次, 防止死循环
        for _ in range(100):
            if current == END:
                trajectory.append("END")
                break
            trajectory.append(current)

            # 调用节点
            node_fn = self.nodes[current]
            update = node_fn(state)
            if isinstance(update, dict):
                for k, v in update.items():
                    state[k] = v

            # 找下一跳
            if (current,) in self.conditional_edges:
                router_fn, mapping = self.conditional_edges[(current,)]
                decision = router_fn(state)
                next_node = mapping.get(decision, END)
            elif (current,) in self.edges:
                next_node = self.edges[(current,)]
            else:
                next_node = END

            current = next_node

        state["__trajectory__"] = trajectory
        return state


class StateGraph:
    def __init__(self, state_type):
        self.state_type = state_type
        self.nodes = {}
        self.edges = {}
        self.conditional_edges = {}
        self.entry_point = None

    def add_node(self, name, fn):
        self.nodes[name] = fn

    def add_edge(self, src, dst):
        self.edges[(src,)] = dst

    def set_entry_point(self, name):
        self.entry_point = name

    def add_conditional_edges(self, src, router_fn, mapping):
        self.conditional_edges[(src,)] = (router_fn, mapping)

    def compile(self):
        return _MiniGraph(self.nodes, self.edges, self.conditional_edges, self.entry_point)


lg_graph = types.ModuleType("langgraph.graph")
lg_graph.END = END
lg_graph.StateGraph = StateGraph
sys.modules["langgraph"] = types.ModuleType("langgraph")
sys.modules["langgraph.graph"] = lg_graph

# 0.5 Mock rich.logging.RichHandler
_rich = types.ModuleType("rich")
_rich_logging = types.ModuleType("rich.logging")
class _RichHandler:
    def __init__(self, *a, **kw):
        self.formatter = None
        self.level = 0  # logging.NOTSET, 接受所有 level
    def setFormatter(self, fmt): self.formatter = fmt
    def setLevel(self, lvl): self.level = lvl
    def handle(self, record): pass
    def emit(self, record): pass
_rich_logging.RichHandler = _RichHandler
sys.modules["rich"] = _rich
sys.modules["rich.logging"] = _rich_logging

# 0.6 Mock dotenv (main.py 用)
class _Dotenv:
    def load_dotenv(self, *a, **kw): pass
sys.modules.setdefault("dotenv", _Dotenv())

# 0.6 mock langchain_openai 在 sandbox 内
sys.modules["dotenv"] = types.ModuleType("dotenv")
sys.modules["dotenv"].load_dotenv = lambda *a, **kw: None

# ============================================================
# Part 1: 现在才 import 项目代码 (mock 已就位)
# ============================================================
print("[Setup] mocks injected, importing project modules...")
from backend.app.pipeline.graph.workflow import build_graph, _should_replan
from backend.app.pipeline.graph.state import GraphState
from backend.app.pipeline.schemas import (StoryAnalysis, Storyboard, PromptPlan,
                         Character, Scene, Shot, DialogueLine,
                         PromptVariant, ShotPrompts,
                         CameraMovement, FramingStyle, MoodTone, VisualStyle,
                         TargetModel)
from backend.app.pipeline.agents import (step1_analyze, step2_storyboard, step3_plan_prompts,
                         step4_write_prompts, step5_consistency_check,
                         step6_model_adapter)
from backend.app.pipeline.utils import get_llm, export_prompts_to_excel
print("[Setup] project modules imported OK")


# ============================================================
# Part 2: Mock LLM —— 按调用顺序返回预设响应
# ============================================================

class MockLLM:
    """按预设队列返回 AIMessage, 队空则抛错。"""

    def __init__(self, responses):
        self.responses = list(responses)  # 复制
        self.call_log = []
        self.cursor = 0

    def invoke(self, messages):
        self.call_log.append([m.content if hasattr(m, "content") else str(m)[:80]
                              for m in messages])
        if self.cursor >= len(self.responses):
            raise RuntimeError(f"MockLLM exhausted at call #{self.cursor + 1}")
        content = self.responses[self.cursor]
        self.cursor += 1
        return AIMessage(content=content)


def _wrap_in_fence(content):
    return "```json\n" + content + "\n```"


# ============================================================
# Part 3: 构造合法的样本 JSON (LLM 模拟输出)
# ============================================================

_SAMPLE_STORY_ANALYSIS_JSON = json.dumps({
    "title": "女巫与龙",
    "genre": "奇幻",
    "tone": "mysterious",
    "visual_style": "cinematic",
    "target_models": ["kling", "jimeng"],
    "characters": [
        {"character_id": "char_001", "name": "艾琳",
         "role": "主角", "appearance": "银袍红发少女, 手持水晶球",
         "personality": "勇敢而好奇", "visual_keywords": ["银袍", "红发", "水晶球"]},
        {"character_id": "char_002", "name": "古龙",
         "role": "导师", "appearance": "金色瞳孔的巨大东方龙, 鳞片泛蓝",
         "personality": "威严而慈祥", "visual_keywords": ["东方龙", "金瞳", "蓝鳞"]},
    ],
    "scenes": [
        {"scene_id": "scene_001", "location": "悬崖古堡外",
         "time_of_day": "深夜", "description": "夜色如墨, 古老的城堡矗立在悬崖之上",
         "characters": ["char_001"], "visual_keywords": ["夜色", "古堡"]},
        {"scene_id": "scene_002", "location": "古堡内部",
         "time_of_day": "室内昏暗", "description": "城堡内部, 封印千年的囚笼",
         "characters": ["char_001", "char_002"], "visual_keywords": ["囚笼", "封印"]},
    ],
    "plot_summary": "年轻女巫艾琳破解城堡封印, 释放出被囚千年的导师——古龙, 二人重逢。",
    "visual_keywords": ["cinematic", "魔法光效", "东方龙"],
}, ensure_ascii=False)


_SAMPLE_STORYBOARD_JSON = json.dumps({
    "title": "女巫与龙",
    "based_on_title": "女巫与龙",
    "shots": [
        {"shot_id": "shot_001", "scene_id": "scene_001", "shot_index": 1,
         "duration_sec": 5.0, "framing": "wide", "camera": "drone_aerial",
         "description": "俯拍, 夜色中的悬崖古堡, 远处蓝光闪烁",
         "characters_in_shot": [], "dialogue": [], "visual_focus": "氛围铺陈",
         "visual_style_override": None},
        {"shot_id": "shot_002", "scene_id": "scene_001", "shot_index": 2,
         "duration_sec": 4.0, "framing": "medium", "camera": "dolly_in",
         "description": "艾琳举起水晶球, 蓝光从球体溢出",
         "characters_in_shot": ["char_001"],
         "dialogue": [{"character_id": "char_001",
                       "line": "封印, 在我的咒语下碎裂吧",
                       "emotion": "坚定", "delivery_type": "dialogue"}],
         "visual_focus": "魔法仪式", "visual_style_override": None},
        {"shot_id": "shot_003", "scene_id": "scene_002", "shot_index": 3,
         "duration_sec": 6.0, "framing": "close_up", "camera": "static",
         "description": "古龙缓缓睁眼, 金瞳与艾琳对视",
         "characters_in_shot": ["char_002"],
         "dialogue": [{"character_id": "char_002",
                       "line": "你终于来了, 继承者",
                       "emotion": "苍老而温和", "delivery_type": "dialogue"}],
         "visual_focus": "情感高潮", "visual_style_override": None},
    ],
    "total_duration_sec": 0.0,
}, ensure_ascii=False)


def _build_prompt_plan_json():
    """动态生成, 含 3 个镜头 x 5 模型 = 15 个 variant。"""
    target_models = ["kling", "jimeng"]
    shot_prompts = []
    for i in range(1, 4):
        variants = [
            {"target_model": m, "prompt_text": "", "negative_prompt": "",
             "aspect_ratio": "16:9", "duration_sec": 4.0 + i, "notes": f"[plan] for {m}"}
            for m in target_models
        ]
        shot_prompts.append({
            "shot_id": f"shot_{i:03d}",
            "variants": variants,
            "dialogue": [],
        })
    return json.dumps({
        "story_title": "女巫与龙",
        "target_models": target_models,
        "shot_prompts": shot_prompts,
    }, ensure_ascii=False)


def _build_writer_response(model, shot_id=""):
    """Step 4 每个变体调一次 LLM, 返回 prompt_text + negative_prompt。

    含模拟时间戳节奏: [0-2s] / [2-4s] ...
    """
    base = f"cinematic {shot_id} scene with magical lighting, high quality, 4K"
    prompt_text = (
        f"{shot_id}\n"
        f"[0-2s] 起始状态: 主体站在古堡前, 蓝光映照\n"
        f"[2-4s] 主体动作: 抬手, 水晶球发光\n"
        f"[4-6s] 环境变化: 封印开始碎裂\n"
        f"镜头节奏: 慢→快\n"
        f"摄影: medium, 50mm, static, cinematic lighting"
    )
    return json.dumps({
        "prompt_text": prompt_text,
        "negative_prompt": "blurry, low quality, watermark",
    }, ensure_ascii=False)


def _build_judge_pass_json():
    """扩展 11 维度 audit 全部 PASS (兼容升级版 _JudgeOutput)。"""
    pass_dim = {"status": "PASS", "issues": []}
    return json.dumps({
        "passed": True,
        "overall_score": 95,
        "confidence": 0.95,
        "issues": [],
        "suggestions": [],
        # 11 维度
        "story_consistency":     pass_dim,
        "character_consistency": pass_dim,
        "scene_consistency":     pass_dim,
        "prop_consistency":      pass_dim,
        "action_consistency":    pass_dim,
        "camera_consistency":    pass_dim,
        "lighting_consistency":  pass_dim,
        "environment_consistency": pass_dim,
        "audio_consistency":     pass_dim,
        "prompt_quality":        pass_dim,
        "negative_prompt":       pass_dim,
        "optimized_prompt":      {"prompt_text": ""},
    }, ensure_ascii=False)


def _build_judge_fail_json():
    """扩展 11 维度 audit, 1 个维度 ERROR (兼容升级版 _JudgeOutput)。"""
    error_dim = {"status": "ERROR", "issues": ["角色 char_002 在 shot_001 中未出场但 prompt 提及"]}
    pass_dim = {"status": "PASS", "issues": []}
    return json.dumps({
        "passed": False,
        "overall_score": 62,
        "confidence": 0.85,
        "issues": ["[character_consistency] 角色 char_002 在 shot_001 中未出场但 prompt 提及"],
        "suggestions": ["重写 shot_001 的 prompt, 删除 char_002 引用"],
        # 11 维度
        "story_consistency":       pass_dim,
        "character_consistency":   error_dim,   # 触发回滚
        "scene_consistency":       pass_dim,
        "prop_consistency":        pass_dim,
        "action_consistency":      pass_dim,
        "camera_consistency":      pass_dim,
        "lighting_consistency":    pass_dim,
        "environment_consistency": pass_dim,
        "audio_consistency":       pass_dim,
        "prompt_quality":          pass_dim,
        "negative_prompt":         pass_dim,
        "optimized_prompt": {"prompt_text": "修正后 Prompt"},
    }, ensure_ascii=False)


# ============================================================
# Part 4: 辅助 —— 初始 state
# ============================================================

def make_initial_state():
    return GraphState(
        story_text="夜色如墨, 古老的城堡矗立在悬崖之上...",
        story_title="女巫与龙",
        max_replans=3,
        replan_count=0,
        story_analysis=None, storyboard=None, prompt_plan=None,
        consistency_report=None, final_outputs=None, messages=[],
    )


# ============================================================
# Part 5: Scenario A —— 顺畅通关
# ============================================================

def scenario_a_smooth_pass():
    print("\n" + "=" * 60)
    print("[Scenario A] 顺畅通关: Step1->6 一次通过")
    print("=" * 60)

    # LLM 调用预算:
    #   Step 1: 1 次 (返回 StoryAnalysis)
    #   Step 2: 1 次 (返回 Storyboard)
    #   Step 3: 1 次 (返回 PromptPlan)
    #   Step 4: 3 镜头 x 5 模型 = 15 次 (每变体一次)
    #   Step 5: 1 次 (judge 通过)
    #   Step 6: 0 次 (纯规则)
    responses = [
        _wrap_in_fence(_SAMPLE_STORY_ANALYSIS_JSON),  # Step 1
        _wrap_in_fence(_SAMPLE_STORYBOARD_JSON),       # Step 2
        _wrap_in_fence(_build_prompt_plan_json()),     # Step 3
    ]
    # Step 4: 3 镜头 x 5 模型 (含 shot_id)
    for i in range(1, 4):
        shot_id = f"shot_{i:03d}"
        for m in ["kling", "jimeng"]:
            responses.append(_wrap_in_fence(_build_writer_response(m, shot_id)))
    # Step 5: 通过
    responses.append(_wrap_in_fence(_build_judge_pass_json()))

    mock = MockLLM(responses)
    # monkey-patch get_llm: 覆盖所有模块级引用
    get_llm.cache_clear()
    import src.utils.llm as llm_mod
    import src.agents._base as base_mod
    import src.agents.step4_writer as step4_mod
    fake = lambda *a, **kw: mock
    llm_mod.get_llm = fake
    base_mod.get_llm = fake
    step4_mod.get_llm = fake

    graph = build_graph()
    state = make_initial_state()
    final = graph.invoke(state)

    # Debug: 查 Excel 落盘
    import os
    out_dir = pathlib.Path("output")
    files = list(out_dir.glob("*.xlsx"))
    print(f"  Excel files in output/: {len(files)} -> {[f.name for f in files]}")
    print(f"  trajectory: {' -> '.join(final['__trajectory__'])}")
    print(f"  final_outputs.summary: {final['final_outputs']['summary']}")

    # ---- Excel 深度校验 ----
    import openpyxl
    xlsx_files = list(pathlib.Path("output").glob("*.xlsx"))
    assert xlsx_files, "Excel 未生成"
    # 取最新文件 (按 mtime)
    xlsx_files.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    wb = openpyxl.load_workbook(xlsx_files[0])
    print(f"  Latest xlsx: {xlsx_files[0].name}")
    # ---- 3 个 Sheet 验证 ----
    assert "Storyboard" in wb.sheetnames
    assert "Prompt Variants" in wb.sheetnames
    assert "演员表" in wb.sheetnames, f"缺 Sheet 3, 当前 sheets: {wb.sheetnames}"

    ws1 = wb["Storyboard"]
    assert ws1.max_row == 4, f"Storyboard 应 1 header + 3 shots = 4 行, 实际 {ws1.max_row}"

    ws2 = wb["Prompt Variants"]
    # 3 镜头 x 5 模型 = 15 行 + 1 header
    assert ws2.max_row == 7, f"Prompt Variants 应 1 header + 6 = 7 行, 实际 {ws2.max_row}"
    cell_b2 = ws2.cell(row=2, column=2).value
    assert cell_b2 == "kling", f"第一行应是 kling, 实际 {cell_b2}"

    # ---- Sheet 3 验证 (演员表) ----
    ws3 = wb["演员表"]
    # 演员表应有角色数据 (至少 1 header + 角色数)
    assert ws3.max_row >= 2, f"演员表应至少有 1 header + 1 角色, 实际 {ws3.max_row}"
    # 检查表头
    assert ws3.cell(row=1, column=1).value == "角色名"
    assert ws3.cell(row=1, column=2).value == "角色定位"
    print(f"  Excel OK: 3 sheets, Storyboard {ws1.max_row}x{ws1.max_column}, "
          f"Variants {ws2.max_row}x{ws2.max_column}, "
          f"演员表 {ws3.max_row}x{ws3.max_column}")

    # 断言
    assert "step6_model_adapter" in final["__trajectory__"], "未到达 step6"
    assert final["__trajectory__"][-1] == "END", "未到达 END"
    assert final["final_outputs"] is not None, "final_outputs 为空"
    assert final["story_analysis"] is not None
    assert final["storyboard"] is not None
    assert final["prompt_plan"] is not None
    assert final["consistency_report"]["passed"] is True
    assert final["replan_count"] == 0, f"replan_count 应为 0, 实际 {final['replan_count']}"
    # final_outputs.by_model 应有 5 个模型各 3 条
    for m, items in final["final_outputs"]["by_model"].items():
        assert len(items) == 3, f"{m} ?? 3 ?, ?? {len(items)}"
        if m == "kling":
            for it in items:
                assert it["prompt_text"], f"Kling ?????????? (shot={it['shot_id']})"
        if m == "jimeng":
            for it in items:
                assert it["prompt_text"], f"Jimeng ?????????? (shot={it['shot_id']})"
    print("  [A] GREEN smooth pass")


# ============================================================
# Part 6: Scenario B —— 回滚重试
# ============================================================

def scenario_b_replan():
    print("\n" + "=" * 60)
    print("[Scenario B] 回滚重试: Step5 fail -> Step3 -> Step5 pass")
    print("=" * 60)

    # LLM 调用预算:
    #   1: 1 (分析)
    #   2: 1 (分镜)
    #   3 (第一次): 1 (规划)
    #   4 (第一次): 15 (写)
    #   5 (第一次): 1 (失败)
    #   ---- 回滚到 Step 3 ----
    #   3 (第二次): 1 (重规划)
    #   4 (第二次): 15 (重写)
    #   5 (第二次): 1 (通过)
    responses = [
        _wrap_in_fence(_SAMPLE_STORY_ANALYSIS_JSON),  # Step 1
        _wrap_in_fence(_SAMPLE_STORYBOARD_JSON),       # Step 2
        _wrap_in_fence(_build_prompt_plan_json()),     # Step 3 (第 1 次)
    ]
    # Step 4 (第 1 次)
    for i in range(1, 4):
        shot_id = f"shot_{i:03d}"
        for m in ["kling", "jimeng"]:
            responses.append(_wrap_in_fence(_build_writer_response(m, shot_id)))
    # Step 5 (第 1 次): 失败
    responses.append(_wrap_in_fence(_build_judge_fail_json()))
    # Step 3 (第 2 次): 重规划
    responses.append(_wrap_in_fence(_build_prompt_plan_json()))
    # Step 4 (第 2 次): 重写
    for i in range(1, 4):
        shot_id = f"shot_{i:03d}"
        for m in ["kling", "jimeng"]:
            responses.append(_wrap_in_fence(_build_writer_response(m, shot_id)))
    # Step 5 (第 2 次): 通过
    responses.append(_wrap_in_fence(_build_judge_pass_json()))

    mock = MockLLM(responses)
    get_llm.cache_clear()
    import src.utils.llm as llm_mod
    import src.agents._base as base_mod
    import src.agents.step4_writer as step4_mod
    fake = lambda *a, **kw: mock
    llm_mod.get_llm = fake
    base_mod.get_llm = fake
    step4_mod.get_llm = fake

    graph = build_graph()
    state = make_initial_state()
    state["max_replans"] = 3
    final = graph.invoke(state)

    # Debug: 查 Excel 落盘
    import os
    out_dir = pathlib.Path("output")
    files = list(out_dir.glob("*.xlsx"))
    print(f"  Excel files in output/: {len(files)} -> {[f.name for f in files]}")
    print(f"  trajectory: {' -> '.join(final['__trajectory__'])}")
    print(f"  replan_count = {final['replan_count']}")

    # 断言
    traj = final["__trajectory__"]
    # Step5 应出现 2 次
    assert traj.count("step5_consistency_check") == 2, \
        f"Step5 应出现 2 次, 实际 {traj.count('step5_consistency_check')}"
    # Step3 应出现 2 次 (首次 + 回滚)
    assert traj.count("step3_plan_prompts") == 2, \
        f"Step3 应出现 2 次, 实际 {traj.count('step3_plan_prompts')}"
    # Step4 应出现 2 次
    assert traj.count("step4_write_prompts") == 2, \
        f"Step4 应出现 2 次, 实际 {traj.count('step4_write_prompts')}"
    # 路由轨迹应包含 step3 出现于 step5 之后 (回滚证据)
    idx_step5_first = traj.index("step5_consistency_check")
    idx_step3_second = traj.index("step3_plan_prompts", idx_step5_first + 1)
    assert idx_step3_second > idx_step5_first, "回滚方向错误"
    # 最终应通过
    assert final["replan_count"] == 1, f"replan_count 应为 1, 实际 {final['replan_count']}"
    assert final["consistency_report"]["passed"] is True
    assert final["final_outputs"] is not None
    # 总 LLM 调用数 = 1+1+1+15+1+1+15+1 = 36
    assert mock.cursor == 18, f"LLM 调用数应为 36, 实际 {mock.cursor}"
    print(f"  LLM call count = {mock.cursor}")
    print("  [B] GREEN replan scenario")


# ============================================================
# Part 7: Scenario C —— 回滚超限
# ============================================================

def scenario_c_max_replans_exceeded():
    print("\n" + "=" * 60)
    print("[Scenario C] 回滚超限: Step5 连续失败, 达到 max_replans 终止")
    print("=" * 60)

    # max_replans = 1, 第二次失败就直接终止
    responses = [
        _wrap_in_fence(_SAMPLE_STORY_ANALYSIS_JSON),
        _wrap_in_fence(_SAMPLE_STORYBOARD_JSON),
        _wrap_in_fence(_build_prompt_plan_json()),
    ]
    for _ in range(3):
        for m in ["kling", "jimeng"]:
            responses.append(_wrap_in_fence(_build_writer_response(m)))
    responses.append(_wrap_in_fence(_build_judge_fail_json()))
    # 回滚后再次失败 -> 终止
    responses.append(_wrap_in_fence(_build_prompt_plan_json()))
    for _ in range(3):
        for m in ["kling", "jimeng"]:
            responses.append(_wrap_in_fence(_build_writer_response(m)))
    responses.append(_wrap_in_fence(_build_judge_fail_json()))

    mock = MockLLM(responses)
    get_llm.cache_clear()
    import src.utils.llm as llm_mod
    import src.agents._base as base_mod
    import src.agents.step4_writer as step4_mod
    fake = lambda *a, **kw: mock
    llm_mod.get_llm = fake
    base_mod.get_llm = fake
    step4_mod.get_llm = fake

    graph = build_graph()
    state = make_initial_state()
    state["max_replans"] = 1
    final = graph.invoke(state)

    # Debug: 查 Excel 落盘
    import os
    out_dir = pathlib.Path("output")
    files = list(out_dir.glob("*.xlsx"))
    print(f"  Excel files in output/: {len(files)} -> {[f.name for f in files]}")
    print(f"  trajectory: {' -> '.join(final['__trajectory__'])}")
    print(f"  replan_count = {final['replan_count']}")

    traj = final["__trajectory__"]
    assert traj[-1] == "END", "应到达 END"
    assert final["replan_count"] == 1, f"replan_count 应为 1, 实际 {final['replan_count']}"
    # step6 不应执行 (因 step6_failed 路由终止)
    # 但 mock 路由: replan_count=1, max=1, replan_count < max 为 False -> step6_failed -> END
    # 实际: step5 fail -> step3 fail -> step5 fail (replan_count=1, 1<1 False) -> step6_failed -> END
    assert final["final_outputs"] is None, "step6 未执行, final_outputs 应为 None"
    print("  [C] GREEN max_replans exceeded")


# ============================================================
# Part 8: 主入口
# ============================================================




# ============================================================
# Part 8: Scenario D —— HITL reject 触发回滚
# ============================================================

def scenario_d_hitl_reject():
    print("\n" + "=" * 60)
    print("[Scenario D] HITL reject: Step4.5 reject -> 回滚 Step3")
    print("=" * 60)

    responses = [
        _wrap_in_fence(_SAMPLE_STORY_ANALYSIS_JSON),  # Step 1
        _wrap_in_fence(_SAMPLE_STORYBOARD_JSON),       # Step 2
        _wrap_in_fence(_build_prompt_plan_json()),     # Step 3 (第 1 次)
    ]
    # Step 4 (第 1 次): 3 镜头 x 5 模型 = 15
    for i in range(1, 4):
        shot_id = f"shot_{i:03d}"
        for m in ["kling", "jimeng"]:
            responses.append(_wrap_in_fence(_build_writer_response(m, shot_id)))
    # ---- reject 触发回滚 ----
    # Step 3 (第 2 次): 重规划
    responses.append(_wrap_in_fence(_build_prompt_plan_json()))
    # Step 4 (第 2 次): 15 次重写
    for i in range(1, 4):
        shot_id = f"shot_{i:03d}"
        for m in ["kling", "jimeng"]:
            responses.append(_wrap_in_fence(_build_writer_response(m, shot_id)))
    # Step 5 (最终通过)
    responses.append(_wrap_in_fence(_build_judge_pass_json()))

    mock = MockLLM(responses)
    get_llm.cache_clear()
    import src.utils.llm as llm_mod
    import src.agents._base as base_mod
    import src.agents.step4_writer as step4_mod
    fake = lambda *a, **kw: mock
    llm_mod.get_llm = fake
    base_mod.get_llm = fake
    step4_mod.get_llm = fake

    graph = build_graph()
    state = make_initial_state()

    # P0 注入: 三个 Review 节点各消费一条反馈 (FIFO)。
    #   - step1_5_review  -> accept
    #   - step2_5_review  -> accept
    #   - step4_5_review  -> reject:Prompt 不满意, 重新规划 (触发回滚 Step3)
    state["human_feedback"] = [
        "accept",
        "accept",
        "reject:Prompt 不满意, 重新规划",
    ]

    final = graph.invoke(state)
    traj = final["__trajectory__"]
    print(f"  trajectory: {' -> '.join(traj)}")
    print(f"  replan_count = {final['replan_count']}")

    # 断言 1: step4_5_review 至少出现 2 次 (1 次 reject + 1 次 accept)
    assert traj.count("step4_5_review") >= 2, \
        f"Step4_5_review 应至少出现 2 次, 实际 {traj.count('step4_5_review')} 次"
    # 断言 2: reject 后应回滚到 step3_plan_prompts
    first_review_idx = traj.index("step4_5_review")
    next_after_review = traj[first_review_idx + 1]
    assert next_after_review == "step3_plan_prompts", \
        f"Step4.5 reject 后应到 step3, 实际 {next_after_review}"
    # 断言 3: replan_count >= 1 (reject 计入一次回滚)
    assert final["replan_count"] >= 1, \
        f"reject 应触发 replan, 实际 replan_count={final['replan_count']}"
    # 断言 4: 整体通关
    assert traj[-1] == "END", "未到达 END"
    assert "step6_model_adapter" in traj, "未到达 step6"
    assert final["final_outputs"] is not None, "final_outputs 为空"
    # 断言 5: step4_5_review_decision 字段最终是 accept (第二次过 review)
    assert final.get("step4_5_review_decision") == "accept", \
        f"最终 step4_5_review_decision 应为 accept, 实际 {final.get('step4_5_review_decision')}"
    print("  [D] GREEN HITL reject scenario")


if __name__ == "__main__":
    print("=" * 60)
    print("E2E Test Suite (sandbox-friendly with mocks)")
    print("=" * 60)

    try:
        scenario_a_smooth_pass()
    except Exception as e:
        import traceback
        print(f"\n[Scenario A] RED: {e}")
        traceback.print_exc()
        sys.exit(1)

    try:
        scenario_b_replan()
    except Exception as e:
        import traceback
        print(f"\n[Scenario B] RED: {e}")
        traceback.print_exc()
        sys.exit(1)

    try:
        scenario_c_max_replans_exceeded()
    except Exception as e:
        import traceback
        print(f"\n[Scenario C] RED: {e}")
        traceback.print_exc()
        sys.exit(1)

    try:
        scenario_d_hitl_reject()
    except Exception as e:
        import traceback
        print(f"\n[Scenario D] RED: {e}")
        traceback.print_exc()
        sys.exit(1)

    print("\n" + "=" * 60)
    print("ALL E2E SCENARIOS GREEN")
    print("=" * 60)
